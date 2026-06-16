import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

st.set_page_config(page_title="Completación de Dispos", page_icon="📦", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .block-container { padding: 2rem 3rem; max-width: 1400px; }
    .app-header {
        background: linear-gradient(135deg, #1a2f4e 0%, #0d4f3c 100%);
        border-radius: 12px; padding: 2rem 2.5rem; margin-bottom: 2rem; color: white;
    }
    .app-header h1 { margin: 0; font-size: 1.8rem; font-weight: 700; letter-spacing: -0.5px; }
    .app-header p  { margin: 0.4rem 0 0; font-size: 0.9rem; opacity: 0.75; }
    .section-title {
        font-size: 0.75rem; font-weight: 700; letter-spacing: 0.12em;
        text-transform: uppercase; color: #6b7280;
        margin: 1.5rem 0 0.75rem; padding-bottom: 0.4rem; border-bottom: 1px solid #e5e7eb;
    }
    .kpi-grid { display: flex; gap: 1rem; margin: 1.2rem 0; flex-wrap: wrap; }
    .kpi-card { flex: 1; min-width: 120px; border-radius: 10px; padding: 1.1rem 1.3rem; border: 1px solid #e5e7eb; }
    .kpi-card.green  { background: #f0fdf4; border-color: #bbf7d0; }
    .kpi-card.yellow { background: #fefce8; border-color: #fde68a; }
    .kpi-card.red    { background: #fff1f2; border-color: #fecdd3; }
    .kpi-card.blue   { background: #eff6ff; border-color: #bfdbfe; }
    .kpi-card.gray   { background: #f9fafb; border-color: #e5e7eb; }
    .kpi-label { font-size: 0.72rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.08em; color: #6b7280; margin-bottom: 0.3rem; }
    .kpi-value { font-size: 1.9rem; font-weight: 700; line-height: 1; color: #111827; }
    .kpi-sub   { font-size: 0.78rem; color: #6b7280; margin-top: 0.25rem; }
    .scenario-tabs .stTabs [data-baseweb="tab"] { font-weight: 600; }
    .tip-box { background: #f8fafc; border-left: 3px solid #3b82f6; border-radius: 0 8px 8px 0; padding: 0.75rem 1rem; font-size: 0.82rem; color: #374151; margin: 0.75rem 0; }
    .warn-box { background: #fffbeb; border-left: 3px solid #f59e0b; border-radius: 0 8px 8px 0; padding: 0.75rem 1rem; font-size: 0.82rem; color: #374151; margin: 0.75rem 0; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="app-header">
    <h1>📦 Completación de Dispos — 3 Escenarios</h1>
    <p>INV · INV+PLAN_DIA1 · INV+PLAN_SEMANAL · Llave: ESTILO_EQ + DTITULAR · Solo líneas ACTIVAS reciben inventario</p>
</div>
""", unsafe_allow_html=True)

# ── Constantes ────────────────────────────────────────────────────────────────
DEFAULT_ENTREGA = {
    '01-EXPEDITE': 1, '02-PAST DUE': 2, '03-DUE': 3, '04-AHEAD': 4, '05-AHEAD': 5,
}
CAPACIDAD_LOTSIZE = [
    {'LOTSIZE': 'D-2200', 'MIX': 'BLEACH', 'LOTES': 14, 'PRIORIDAD': 1, 'ACTIVO': True},
    {'LOTSIZE': 'F-1000', 'MIX': 'BLEACH', 'LOTES':  7, 'PRIORIDAD': 2, 'ACTIVO': True},
    {'LOTSIZE': 'A-4000', 'MIX': 'DYE',    'LOTES':  4, 'PRIORIDAD': 1, 'ACTIVO': True},
    {'LOTSIZE': 'B-3300', 'MIX': 'DYE',    'LOTES':  8, 'PRIORIDAD': 2, 'ACTIVO': True},
    {'LOTSIZE': 'C-2600', 'MIX': 'DYE',    'LOTES': 38, 'PRIORIDAD': 3, 'ACTIVO': True},
    {'LOTSIZE': 'D-2200', 'MIX': 'DYE',    'LOTES': 29, 'PRIORIDAD': 4, 'ACTIVO': True},
    {'LOTSIZE': 'F-1000', 'MIX': 'DYE',    'LOTES': 33, 'PRIORIDAD': 5, 'ACTIVO': True},
]
CASCADA_COLOR_NOMBRES = {
    '0': 'BLANCO', '1': 'AMARILLO', '2': 'NARANJA', '3': 'ROJO',
    '4': 'MORADO', '5': 'ROYAL',    '6': 'VERDE',   '7': 'CAFÉ',
    '8': 'GRIS',   '9': 'NEGRO',
}
ESCENARIOS = [
    {'key': 'INV',      'label': '📦 Solo INV',           'col_inv': 'INV',  'col_extra': None},
    {'key': 'DIA1',     'label': '📅 INV + PLAN_DIA1',    'col_inv': 'INV',  'col_extra': 'PLAN_INS_DIA1'},
    {'key': 'SEMANAL',  'label': '📆 INV + PLAN_SEMANAL', 'col_inv': 'INV',  'col_extra': 'PLAN_INS'},
]

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuración")

    st.markdown('<p class="section-title">Prioridad ENTREGA</p>', unsafe_allow_html=True)
    st.markdown('<div class="tip-box">Menor número = mayor prioridad.</div>', unsafe_allow_html=True)
    entrega_pesos = {}
    for val, default in DEFAULT_ENTREGA.items():
        entrega_pesos[val] = st.number_input(val, min_value=1, max_value=99,
                                             value=default, key=f"e_{val}")

    st.markdown('<p class="section-title">Capacidad LOTSIZE + MIX</p>', unsafe_allow_html=True)
    st.markdown('<div class="tip-box">Cuántas DISPOs puede procesar cada combinación. 🔴 = no recibe asignación.</div>',
                unsafe_allow_html=True)
    capacidad_cfg = []
    for i, row in enumerate(CAPACIDAD_LOTSIZE):
        col_a, col_b, col_c = st.columns([2, 1, 1])
        with col_a:
            st.markdown(f"**{row['LOTSIZE']} · {row['MIX']}** (P{row['PRIORIDAD']})")
        with col_b:
            lotes = st.number_input('Lotes', min_value=0, max_value=9999,
                                    value=row['LOTES'], key=f"lotes_{i}",
                                    label_visibility='collapsed')
        with col_c:
            activo_ls = st.selectbox('Estado', options=[True, False],
                                     index=0 if row['ACTIVO'] else 1,
                                     format_func=lambda x: '🟢' if x else '🔴',
                                     key=f"activo_ls_{i}",
                                     label_visibility='collapsed')
        capacidad_cfg.append({**row, 'LOTES': lotes, 'ACTIVO': activo_ls})

    st.markdown('<p class="section-title">Cascada de color (COLOR_A dígito 3)</p>', unsafe_allow_html=True)
    st.markdown('<div class="warn-box">🔴 INACTIVO = no recibe inventario.</div>', unsafe_allow_html=True)
    cascada_activo = {}
    for digito, nombre in CASCADA_COLOR_NOMBRES.items():
        cascada_activo[digito] = st.selectbox(
            f"{digito} — {nombre}", options=[1, 0], index=0,
            format_func=lambda x: "🟢 ACTIVO" if x == 1 else "🔴 INACTIVO",
            key=f"color_{digito}"
        )

# ── Funciones core ────────────────────────────────────────────────────────────
def leer_excel(file):
    for header_row in range(10):
        try:
            df = pd.read_excel(file, header=header_row)
            df.columns = df.columns.str.strip().str.upper()
            if 'DISPO' in df.columns and 'ESTILO_EQ' in df.columns:
                return df, header_row
        except Exception:
            continue
    return None, None


def extraer_digito3(color_a):
    try:
        s = str(int(float(color_a))) if str(color_a).replace('.', '').isdigit() else str(color_a)
        return s[2] if len(s) >= 3 else None
    except Exception:
        return None


def procesar(df, entrega_pesos, cascada_activo, col_extra=None, capacidad_cfg=None):
    df = df.copy()
    df.columns = df.columns.str.strip().str.upper()
    df = df.dropna(subset=['ESTILO_EQ', 'DTITULAR', 'LBS_C', 'INV']).reset_index(drop=True)

    # Inventario efectivo según escenario
    if col_extra and col_extra in df.columns:
        df['INV_EFECTIVO'] = df['INV'].fillna(0) + df[col_extra].fillna(0)
    else:
        df['INV_EFECTIVO'] = df['INV'].fillna(0)

    # Cascada de color
    if 'COLOR_A' in df.columns:
        df['_DIGITO3']     = df['COLOR_A'].apply(extraer_digito3)
        df['COLOR_NOMBRE'] = df['_DIGITO3'].map(CASCADA_COLOR_NOMBRES).fillna('DESCONOCIDO')
        df['ACTIVO']       = df['_DIGITO3'].map(cascada_activo).fillna(0).astype(int)
    else:
        df['_DIGITO3']     = None
        df['COLOR_NOMBRE'] = 'SIN COLOR_A'
        df['ACTIVO']       = 1

    # Prioridad
    df['_PESO_E'] = df['ENTREGA'].map(entrega_pesos).fillna(99).astype(int)
    cap_cfg = capacidad_cfg or CAPACIDAD_LOTSIZE
    ls_mix_prio = {(r['LOTSIZE'], r['MIX']): r['PRIORIDAD'] for r in cap_cfg}
    if 'LOTSIZE' in df.columns and 'MIX' in df.columns:
        df['_PESO_L'] = df.apply(lambda r: ls_mix_prio.get((r['LOTSIZE'], r['MIX']), 99), axis=1).astype(int)
    else:
        df['_PESO_L'] = 99
    df['_ORDEN'] = df['_PESO_E'] * 100 + df['_PESO_L']
    df['_IDX']   = range(len(df))

    # Capacidad dict
    cap_dict = {(r['LOTSIZE'], r['MIX']): {'lotes': r['LOTES'], 'activo': r['ACTIVO']} for r in cap_cfg}

    def ls_mix_activo(row):
        if 'LOTSIZE' not in row.index or 'MIX' not in row.index:
            return True
        cfg = cap_dict.get((row.get('LOTSIZE'), row.get('MIX')))
        return True if cfg is None else cfg['activo']

    df['_LS_ACTIVO'] = df.apply(ls_mix_activo, axis=1)

    df_sorted = df.sort_values(
        ['ESTILO_EQ', 'DTITULAR', 'ACTIVO', '_ORDEN', '_IDX'],
        ascending=[True, True, False, True, True]
    )

    lotes_usados     = {k: 0 for k in cap_dict}
    lbs_asignado     = [0.0] * len(df_sorted)
    inv_restante_col = [0.0] * len(df_sorted)
    idx_map          = {idx: pos for pos, idx in enumerate(df_sorted.index)}

    # Pasada 1 — respeta capacidad
    for (esq, dtit), grupo in df_sorted.groupby(['ESTILO_EQ', 'DTITULAR'], sort=False):
        inv_disp  = grupo['INV_EFECTIVO'].iloc[0]
        acumulado = 0
        for orig_idx, fila in grupo.iterrows():
            pos          = idx_map[orig_idx]
            lbs_c        = fila['LBS_C']
            activo_color = fila['ACTIVO']
            activo_ls    = fila['_LS_ACTIVO']

            if activo_color == 0 or not activo_ls:
                lbs_asignado[pos]     = 0
                inv_restante_col[pos] = max(0, inv_disp - acumulado)
                continue

            key = (fila.get('LOTSIZE'), fila.get('MIX')) if 'LOTSIZE' in fila.index and 'MIX' in fila.index else None
            cfg = cap_dict.get(key) if key else None

            if cfg is not None and lotes_usados.get(key, 0) >= cfg['lotes']:
                lbs_asignado[pos]     = -1   # sentinel pasada 2
                inv_restante_col[pos] = max(0, inv_disp - acumulado)
                continue

            disponible = max(0, inv_disp - acumulado)
            asignado   = min(lbs_c, disponible)
            acumulado += lbs_c
            lbs_asignado[pos]     = asignado
            inv_restante_col[pos] = max(0, inv_disp - acumulado)
            if key and key in lotes_usados:
                lotes_usados[key] += 1

    # Pasada 2 — si todos los LOTSIZE activos llegaron a su límite
    todos_cubiertos = all(
        lotes_usados.get(k, 0) >= v['lotes']
        for k, v in cap_dict.items() if v['activo']
    )
    if todos_cubiertos:
        for (esq, dtit), grupo in df_sorted.groupby(['ESTILO_EQ', 'DTITULAR'], sort=False):
            inv_disp  = grupo['INV_EFECTIVO'].iloc[0]
            acumulado = sum(lbs_asignado[idx_map[i]] for i in grupo.index if lbs_asignado[idx_map[i]] > 0)
            for orig_idx, fila in grupo.iterrows():
                pos = idx_map[orig_idx]
                if lbs_asignado[pos] != -1:
                    continue
                lbs_c      = fila['LBS_C']
                disponible = max(0, inv_disp - acumulado)
                asignado   = min(lbs_c, disponible)
                acumulado += lbs_c
                lbs_asignado[pos]     = asignado
                inv_restante_col[pos] = max(0, inv_disp - acumulado)
    else:
        lbs_asignado = [0 if v == -1 else v for v in lbs_asignado]

    df_sorted['LBS_ASIGNADO'] = lbs_asignado
    df_sorted['INV_RESTANTE'] = inv_restante_col
    df_sorted['LBS_FALTANTE'] = (df_sorted['LBS_C'] - df_sorted['LBS_ASIGNADO']).clip(lower=0)
    df_sorted['PCT_LINEA']    = (df_sorted['LBS_ASIGNADO'] / df_sorted['LBS_C'].replace(0, np.nan)).fillna(0)

    activas_mask = df_sorted['ACTIVO'] == 1
    pct_min      = df_sorted[activas_mask].groupby('DISPO')['PCT_LINEA'].min().rename('_min_pct')
    df_sorted    = df_sorted.merge(pct_min, on='DISPO', how='left')
    df_sorted['_min_pct'] = df_sorted['_min_pct'].fillna(-1)

    def status_dispo(row):
        if row['ACTIVO'] == 0:    return '⛔ INACTIVA'
        p = row['_min_pct']
        if p < 0:  return '⛔ INACTIVA'
        if p >= 1: return '✅ COMPLETA'
        if p > 0:  return '⚠️ PARCIAL'
        return '❌ SIN INVENTARIO'

    df_sorted['STATUS_DISPO'] = df_sorted.apply(status_dispo, axis=1)

    pct_dispo = (
        df_sorted[activas_mask].groupby('DISPO')['LBS_ASIGNADO'].sum() /
        df_sorted[activas_mask].groupby('DISPO')['LBS_C'].sum().replace(0, np.nan)
    ).fillna(0).rename('PCT_DISPO')
    df_sorted = df_sorted.merge(pct_dispo, on='DISPO', how='left')
    df_sorted['PCT_DISPO'] = df_sorted['PCT_DISPO'].fillna(0)

    drop = ['_PESO_E', '_PESO_L', '_ORDEN', '_IDX', '_DIGITO3', '_min_pct', '_LS_ACTIVO']
    return df_sorted.sort_values('_IDX').drop(columns=drop)


def kpis(df_r):
    activas   = df_r[df_r['ACTIVO'] == 1]
    dispo_min = activas.groupby('DISPO')['PCT_LINEA'].min()
    total_d   = dispo_min.count()
    return {
        'total':     total_d,
        'completas': (dispo_min >= 1).sum(),
        'parciales': ((dispo_min > 0) & (dispo_min < 1)).sum(),
        'sin_inv':   (dispo_min == 0).sum(),
        'inactivas': (df_r['ACTIVO'] == 0).sum(),
        'cob':       activas['LBS_ASIGNADO'].sum() / activas['LBS_C'].sum() if len(activas) else 0,
    }


# ── Excel helpers ─────────────────────────────────────────────────────────────
COLOR_H = {'orig': '2F4F7F', 'new': '1D6A40', 'cfg': '7B3F00',
           'sc1': '1E40AF', 'sc2': '065F46', 'sc3': '6B21A8'}
COLOR_ROW = {'ok': 'C6EFCE', 'warn': 'FFEB9C', 'err': 'FFC7CE', 'inac': 'E5E7EB', 'total': 'D1FAE5'}

def make_styles():
    thin = Side(style='thin', color='CCCCCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        'FW':  Font(color='FFFFFF', bold=True, name='Arial', size=10),
        'FN':  Font(name='Arial', size=9),
        'FB':  Font(name='Arial', size=9, bold=True),
        'FI':  Font(name='Arial', size=9, color='9CA3AF', italic=True),
        'FT':  Font(name='Arial', size=9, bold=True, color='065F46'),
        'brd': brd,
    }


def formato_detalle(ws, df_out, n_orig, st_):
    FW, FN, FI, brd = st_['FW'], st_['FN'], st_['FI'], st_['brd']
    for col_idx in range(1, len(df_out.columns) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = PatternFill('solid', start_color=COLOR_H['new'] if col_idx > n_orig else COLOR_H['orig'])
        c.font = FW
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = brd
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    s_idx  = df_out.columns.get_loc('STATUS_DISPO') + 1
    pl_idx = df_out.columns.get_loc('PCT_LINEA') + 1
    pd_idx = df_out.columns.get_loc('PCT_DISPO') + 1
    f_ok   = PatternFill('solid', start_color=COLOR_ROW['ok'])
    f_warn = PatternFill('solid', start_color=COLOR_ROW['warn'])
    f_err  = PatternFill('solid', start_color=COLOR_ROW['err'])
    f_inac = PatternFill('solid', start_color=COLOR_ROW['inac'])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status  = row[s_idx - 1].value
        is_inac = status == '⛔ INACTIVA'
        fill    = f_inac if is_inac else f_ok if status == '✅ COMPLETA' else f_warn if status == '⚠️ PARCIAL' else f_err
        for cell in row:
            cell.font   = FI if is_inac else FN
            cell.border = brd
            if cell.column > n_orig:
                cell.fill      = fill
                cell.alignment = Alignment(horizontal='center')
        row[pl_idx - 1].number_format = '0.0%'
        row[pd_idx - 1].number_format = '0.0%'

    for i, col_name in enumerate(df_out.columns, 1):
        w = (18 if col_name in ('DISPO','ESTILO_EQ','ENTREGA','LOTSIZE','STATUS_DISPO','COLOR_NOMBRE','ESCENARIO')
             else 14 if col_name in ('LBS_C','LBS_ASIGNADO','LBS_FALTANTE','INV','INV_EFECTIVO','INV_RESTANTE','PLAN_INS_DIA1','PLAN_INS')
             else 12 if col_name in ('PCT_LINEA','PCT_DISPO')
             else  9 if col_name == 'ACTIVO' else 11)
        ws.column_dimensions[get_column_letter(i)].width = w


def write_table(ws, title, df, start_row, start_col, header_color, st_):
    FW, FN, FB, FT, brd = st_['FW'], st_['FN'], st_['FB'], st_['FT'], st_['brd']
    end_col = start_col + len(df.columns) - 1
    tc = ws.cell(row=start_row, column=start_col, value=title)
    tc.font = FW; tc.fill = PatternFill('solid', start_color=header_color)
    tc.alignment = Alignment(horizontal='center', vertical='center'); tc.border = brd
    if end_col > start_col:
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    for ci, col_name in enumerate(df.columns, start=start_col):
        c = ws.cell(row=start_row+1, column=ci, value=col_name)
        c.font = FB; c.fill = PatternFill('solid', start_color='D9D9D9')
        c.alignment = Alignment(horizontal='center'); c.border = brd
    for ri, (_, row_data) in enumerate(df.iterrows(), start=start_row+2):
        is_total = any(str(v).upper() == 'TOTAL' for v in row_data.values)
        for ci, val in enumerate(row_data.values, start=start_col):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = brd
            if is_total:
                cell.font = FT; cell.fill = PatternFill('solid', start_color=COLOR_ROW['total'])
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.font = FN
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.0'
                    cell.alignment = Alignment(horizontal='right')
    return start_row + len(df) + 3


def generar_reporte_comparativo(ws, resultados, st_):
    """
    Escribe 4 cuadros, cada uno con 3 columnas de LBS_C (una por escenario).
    resultados = [{'label': ..., 'df': ...}, ...]
    """
    labels = [r['label'] for r in resultados]

    def cuadro_comparativo(ws, title, agg_col, group_col, start_row, start_col, header_color):
        FW, FN, FB, FT, brd = st_['FW'], st_['FN'], st_['FB'], st_['FT'], st_['brd']

        # Construir tabla comparativa
        dfs = []
        for r in resultados:
            completas = r['df'][(r['df']['STATUS_DISPO'] == '✅ COMPLETA') & (r['df']['ACTIVO'] == 1)]
            if group_col in completas.columns:
                t = completas.groupby(group_col)['LBS_C'].sum().rename(r['label'])
            else:
                t = pd.Series(dtype=float, name=r['label'])
            dfs.append(t)

        comp = pd.concat(dfs, axis=1).fillna(0).reset_index()
        comp.columns = [group_col] + labels
        comp = comp.sort_values(labels[0], ascending=False)
        total_row = {group_col: 'TOTAL'}
        for lbl in labels:
            total_row[lbl] = comp[lbl].sum()
        comp = pd.concat([comp, pd.DataFrame([total_row])], ignore_index=True)

        # Escribir
        n_cols = len(comp.columns)
        end_col = start_col + n_cols - 1
        tc = ws.cell(row=start_row, column=start_col, value=title)
        tc.font = FW; tc.fill = PatternFill('solid', start_color=header_color)
        tc.alignment = Alignment(horizontal='center'); tc.border = brd
        if end_col > start_col:
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)

        sc_colors = [COLOR_H['sc1'], COLOR_H['sc2'], COLOR_H['sc3']]
        for ci, col_name in enumerate(comp.columns, start=start_col):
            c = ws.cell(row=start_row+1, column=ci, value=col_name)
            c.font = FW if ci > start_col else FB
            c.fill = PatternFill('solid', start_color=sc_colors[ci - start_col - 1] if ci > start_col else 'D9D9D9')
            c.alignment = Alignment(horizontal='center'); c.border = brd

        for ri, (_, row_data) in enumerate(comp.iterrows(), start=start_row+2):
            is_total = str(row_data.iloc[0]).upper() == 'TOTAL'
            for ci, val in enumerate(row_data.values, start=start_col):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = brd
                if is_total:
                    cell.font = FT; cell.fill = PatternFill('solid', start_color=COLOR_ROW['total'])
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.font = FN
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0.0'
                        cell.alignment = Alignment(horizontal='right')
        return start_row + len(comp) + 3

    # Cuadro 4 — detalle por ENTREGA+LOTSIZE+DISPO+ESTILO_EQ+DTITULAR
    def cuadro4(ws, start_row, start_col):
        FW, FN, FB, FT, brd = st_['FW'], st_['FN'], st_['FB'], st_['FT'], st_['brd']
        group_cols = ['ENTREGA', 'LOTSIZE', 'DISPO', 'ESTILO_EQ', 'DTITULAR']
        dfs = []
        for r in resultados:
            sub = r['df'][(r['df']['ACTIVO'] == 1) & (r['df']['STATUS_DISPO'] == '✅ COMPLETA')]
            gc  = [c for c in group_cols if c in sub.columns]
            t   = sub.groupby(gc, dropna=False)['LBS_C'].sum().rename(r['label']).reset_index()
            dfs.append(t.set_index(gc))
        comp = pd.concat(dfs, axis=1).fillna(0).reset_index()
        cols_out = [c for c in group_cols if c in comp.columns] + labels
        comp = comp[cols_out].sort_values(['ENTREGA', labels[0]], ascending=[True, False])

        title = '📋 Cuadro 4 — Completas por ENTREGA / LOTSIZE / DISPO / ESTILO / TITULAR'
        n_cols  = len(comp.columns)
        end_col = start_col + n_cols - 1
        tc = ws.cell(row=start_row, column=start_col, value=title)
        tc.font = FW; tc.fill = PatternFill('solid', start_color=COLOR_H['cfg'])
        tc.alignment = Alignment(horizontal='center'); tc.border = brd
        if end_col > start_col:
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)

        sc_colors = [COLOR_H['sc1'], COLOR_H['sc2'], COLOR_H['sc3']]
        dim_cols  = [c for c in group_cols if c in comp.columns]
        for ci, col_name in enumerate(comp.columns, start=start_col):
            c = ws.cell(row=start_row+1, column=ci, value=col_name)
            is_sc = col_name in labels
            idx_sc = labels.index(col_name) if is_sc else -1
            c.font = FW if is_sc else FB
            c.fill = PatternFill('solid', start_color=sc_colors[idx_sc] if is_sc else 'D9D9D9')
            c.alignment = Alignment(horizontal='center'); c.border = brd

        for ri, (_, row_data) in enumerate(comp.iterrows(), start=start_row+2):
            for ci, val in enumerate(row_data.values, start=start_col):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = brd; cell.font = FN
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.0'
                    cell.alignment = Alignment(horizontal='right')

        for col in ['E','F','G','H','I','J','K','L','M']:
            ws.column_dimensions[col].width = 16

    # ── Posiciones ────────────────────────────────────────────────────────────
    next_row = cuadro_comparativo(ws, '📋 Cuadro 1 — Completas por ENTREGA',
                                   'LBS_C', 'ENTREGA', 1, 1, COLOR_H['orig'])
    next_row = cuadro_comparativo(ws, '📋 Cuadro 2 — Completas por LOTSIZE',
                                   'LBS_C', 'LOTSIZE', next_row, 1, COLOR_H['orig'])
    cuadro_comparativo(ws, '📋 Cuadro 3 — Completas por COLOR',
                       'LBS_C', 'COLOR_NOMBRE', next_row, 1, COLOR_H['orig'])
    cuadro4(ws, 1, 6)

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18


def generar_excel(resultados, entrega_pesos, cascada_activo, capacidad_cfg):
    st_ = make_styles()
    FW, FN, FB, brd = st_['FW'], st_['FN'], st_['FB'], st_['brd']

    new_cols  = ['COLOR_NOMBRE', 'ACTIVO', 'INV_EFECTIVO',
                 'LBS_ASIGNADO', 'LBS_FALTANTE', 'INV_RESTANTE',
                 'PCT_LINEA', 'PCT_DISPO', 'STATUS_DISPO']

    # Combinar los 3 escenarios en una sola hoja
    frames = []
    for r in resultados:
        df_sc = r['df'].copy()
        df_sc.insert(0, 'ESCENARIO', r['label'])
        frames.append(df_sc)
    df_all = pd.concat(frames, ignore_index=True)

    orig_cols = [c for c in df_all.columns if c not in new_cols and c != 'ESCENARIO']
    col_order = ['ESCENARIO'] + orig_cols + [c for c in new_cols if c in df_all.columns]
    df_out    = df_all[col_order].copy()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name='DETALLE', index=False)
        pd.DataFrame().to_excel(writer, sheet_name='REPORTE', index=False)

        # RESUMEN por escenario
        resumen_rows = []
        for r in resultados:
            k = kpis(r['df'])
            resumen_rows.append({
                'Escenario':            r['label'],
                'DISPOs totales':       k['total'],
                '✅ Completas':         k['completas'],
                '⚠️ Parciales':        k['parciales'],
                '❌ Sin inventario':    k['sin_inv'],
                '⛔ Líneas inactivas':  k['inactivas'],
                '% Cobertura':          f"{k['cob']:.1%}",
            })
        pd.DataFrame(resumen_rows).to_excel(writer, sheet_name='RESUMEN', index=False)

        cfg_e  = pd.DataFrame(list(entrega_pesos.items()), columns=['ENTREGA', 'PESO'])
        cfg_ls = pd.DataFrame(capacidad_cfg or CAPACIDAD_LOTSIZE)
        cfg_cascada = pd.DataFrame([
            {'Dígito': d, 'Color': CASCADA_COLOR_NOMBRES[d],
             'Estado': '🟢 ACTIVO' if cascada_activo.get(d, 1) == 1 else '🔴 INACTIVO'}
            for d in CASCADA_COLOR_NOMBRES
        ])
        cfg_e.to_excel(writer,     sheet_name='CONFIG', index=False, startrow=1, startcol=0)
        cfg_ls.to_excel(writer,    sheet_name='CONFIG', index=False, startrow=1, startcol=3)
        cfg_cascada.to_excel(writer, sheet_name='CONFIG', index=False, startrow=1, startcol=9)

    buf.seek(0)
    wb = load_workbook(buf)

    # ── Formatear DETALLE ─────────────────────────────────────────────────────
    ws    = wb['DETALLE']
    n_orig = 1 + len(orig_cols)   # ESCENARIO + orig cols
    formato_detalle(ws, df_out, n_orig, st_)

    # Color encabezado ESCENARIO
    ws.cell(row=1, column=1).fill = PatternFill('solid', start_color=COLOR_H['cfg'])

    # ── Formatear RESUMEN ─────────────────────────────────────────────────────
    ws_r = wb['RESUMEN']
    sc_fills = [PatternFill('solid', start_color=c) for c in [COLOR_H['sc1'], COLOR_H['sc2'], COLOR_H['sc3']]]
    for col in range(1, 9):
        ws_r.column_dimensions[get_column_letter(col)].width = 22
    for row in ws_r.iter_rows(min_row=1, max_row=ws_r.max_row):
        for cell in row:
            cell.border = brd
            if cell.row == 1:
                cell.font = FW; cell.fill = PatternFill('solid', start_color=COLOR_H['orig'])
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.font = FN
                sc_idx = cell.row - 2
                if cell.column == 1 and 0 <= sc_idx < len(sc_fills):
                    cell.fill = sc_fills[sc_idx]
                    cell.font = Font(color='FFFFFF', bold=True, name='Arial', size=9)

    # ── Formatear CONFIG ──────────────────────────────────────────────────────
    ws_c = wb['CONFIG']
    titulos = {1: 'PRIORIDAD ENTREGA', 4: 'CAPACIDAD LOTSIZE+MIX', 10: 'CASCADA DE COLOR'}
    for col_start, titulo in titulos.items():
        cell = ws_c.cell(row=1, column=col_start, value=titulo)
        cell.font = FW; cell.fill = PatternFill('solid', start_color=COLOR_H['cfg'])
        cell.alignment = Alignment(horizontal='center')
    ws_c.merge_cells('A1:B1'); ws_c.merge_cells('D1:H1'); ws_c.merge_cells('J1:L1')
    for col in ['A','B','D','E','F','G','H','J','K','L']:
        ws_c.column_dimensions[col].width = 16
    for col in ['C','I']:
        ws_c.column_dimensions[col].width = 4
    for row in ws_c.iter_rows(min_row=2, max_row=ws_c.max_row):
        for cell in row:
            cell.font = FB if cell.row == 2 else FN
            cell.border = brd
            if cell.row == 2:
                cell.fill = PatternFill('solid', start_color='D9D9D9')

    # ── REPORTE comparativo ───────────────────────────────────────────────────
    ws_rep = wb['REPORTE']
    generar_reporte_comparativo(ws_rep, resultados, st_)

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out


# ── Layout principal ──────────────────────────────────────────────────────────
col_upload, col_main = st.columns([1, 2.5], gap="large")

with col_upload:
    st.markdown('<p class="section-title">Archivo de entrada</p>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Sube tu archivo Excel", type=['xlsx','xls'],
                                label_visibility="collapsed")

    if uploaded:
        df_raw, header_row = leer_excel(uploaded)
        if df_raw is None:
            st.error("No se encontraron los encabezados. Verifica el archivo.")
        else:
            tiene_color   = 'COLOR_A'        in df_raw.columns
            tiene_dia1    = 'PLAN_INS_DIA1'  in df_raw.columns
            tiene_semanal = 'PLAN_INS'       in df_raw.columns

            st.success(f"**{uploaded.name}**")
            st.caption(f"{len(df_raw):,} filas · {df_raw['DISPO'].nunique():,} dispos · fila {header_row+1}")

            if not tiene_color:
                st.markdown('<div class="warn-box">⚠️ Sin COLOR_A — todas activas.</div>',
                            unsafe_allow_html=True)
            if not tiene_dia1:
                st.markdown('<div class="warn-box">⚠️ Sin PLAN_INS_DIA1 — Escenario 2 = Solo INV.</div>',
                            unsafe_allow_html=True)
            if not tiene_semanal:
                st.markdown('<div class="warn-box">⚠️ Sin PLAN_INS — Escenario 3 = Solo INV.</div>',
                            unsafe_allow_html=True)

            cols_faltantes = [c for c in ['ENTREGA','DISPO','ESTILO_EQ','DTITULAR','LBS_C','INV']
                              if c not in df_raw.columns]
            if cols_faltantes:
                st.warning(f"Columnas faltantes: {cols_faltantes}")
            else:
                preview_cols = [c for c in ['DISPO','ENTREGA','ESTILO_EQ','DTITULAR',
                                            'COLOR_A','LBS_C','INV','PLAN_INS_DIA1','PLAN_INS']
                                if c in df_raw.columns]
                st.markdown('<p class="section-title">Vista previa</p>', unsafe_allow_html=True)
                st.dataframe(df_raw[preview_cols].head(8), use_container_width=True, hide_index=True)

                if st.button("▶ Procesar 3 escenarios", type="primary", use_container_width=True):
                    with st.spinner("Calculando 3 escenarios..."):
                        resultados = []
                        for esc in ESCENARIOS:
                            df_sc = procesar(df_raw, entrega_pesos, cascada_activo,
                                             col_extra=esc['col_extra'],
                                             capacidad_cfg=capacidad_cfg)
                            resultados.append({'key': esc['key'], 'label': esc['label'], 'df': df_sc})
                        st.session_state['resultados'] = resultados

with col_main:
    if 'resultados' in st.session_state:
        resultados = st.session_state['resultados']

        # ── KPIs comparativos ─────────────────────────────────────────────────
        st.markdown('<p class="section-title">Comparativo de escenarios</p>', unsafe_allow_html=True)
        kpi_cols = st.columns(3)
        sc_colors_bg = ['#eff6ff', '#f0fdf4', '#faf5ff']
        sc_colors_bd = ['#bfdbfe', '#bbf7d0', '#e9d5ff']
        for i, r in enumerate(resultados):
            k = kpis(r['df'])
            with kpi_cols[i]:
                st.markdown(f"""
                <div style="background:{sc_colors_bg[i]};border:1px solid {sc_colors_bd[i]};
                            border-radius:10px;padding:1rem 1.2rem;margin-bottom:0.5rem;">
                  <div style="font-size:0.7rem;font-weight:700;text-transform:uppercase;
                              letter-spacing:0.1em;color:#6b7280;margin-bottom:0.5rem;">{r['label']}</div>
                  <div style="display:flex;justify-content:space-between;margin-bottom:0.3rem;">
                    <span style="font-size:0.8rem;">✅ Completas</span>
                    <strong>{k['completas']:,}</strong>
                  </div>
                  <div style="display:flex;justify-content:space-between;margin-bottom:0.3rem;">
                    <span style="font-size:0.8rem;">⚠️ Parciales</span>
                    <strong>{k['parciales']:,}</strong>
                  </div>
                  <div style="display:flex;justify-content:space-between;margin-bottom:0.3rem;">
                    <span style="font-size:0.8rem;">❌ Sin inv</span>
                    <strong>{k['sin_inv']:,}</strong>
                  </div>
                  <div style="display:flex;justify-content:space-between;border-top:1px solid {sc_colors_bd[i]};
                              padding-top:0.4rem;margin-top:0.4rem;">
                    <span style="font-size:0.8rem;">📦 Cobertura</span>
                    <strong>{k['cob']:.1%}</strong>
                  </div>
                </div>
                """, unsafe_allow_html=True)

        # ── Tabla detalle con tabs por escenario ──────────────────────────────
        st.markdown('<p class="section-title">Detalle por escenario</p>', unsafe_allow_html=True)
        tabs = st.tabs([r['label'] for r in resultados])

        cols_show_base = ['DISPO','ENTREGA','LOTSIZE','ESTILO_EQ','DTITULAR',
                          'COLOR_NOMBRE','ACTIVO','MIX',
                          'LBS_C','INV','INV_EFECTIVO',
                          'LBS_ASIGNADO','LBS_FALTANTE',
                          'PCT_LINEA','PCT_DISPO','STATUS_DISPO']

        for tab, r in zip(tabs, resultados):
            with tab:
                df_r = r['df']
                fc1, fc2, fc3 = st.columns(3)
                with fc1:
                    filtro_status = st.multiselect(
                        "Status", ['✅ COMPLETA','⚠️ PARCIAL','❌ SIN INVENTARIO','⛔ INACTIVA'],
                        default=['✅ COMPLETA','⚠️ PARCIAL','❌ SIN INVENTARIO','⛔ INACTIVA'],
                        key=f"fs_{r['key']}", label_visibility="collapsed"
                    )
                with fc2:
                    filtro_entrega = st.multiselect(
                        "ENTREGA", sorted(df_r['ENTREGA'].dropna().unique()),
                        default=sorted(df_r['ENTREGA'].dropna().unique()),
                        key=f"fe_{r['key']}", label_visibility="collapsed"
                    )
                with fc3:
                    if 'COLOR_NOMBRE' in df_r.columns:
                        filtro_color = st.multiselect(
                            "Color", sorted(df_r['COLOR_NOMBRE'].dropna().unique()),
                            default=sorted(df_r['COLOR_NOMBRE'].dropna().unique()),
                            key=f"fc_{r['key']}", label_visibility="collapsed"
                        )
                    else:
                        filtro_color = None

                df_vis = df_r[df_r['STATUS_DISPO'].isin(filtro_status) & df_r['ENTREGA'].isin(filtro_entrega)]
                if filtro_color and 'COLOR_NOMBRE' in df_r.columns:
                    df_vis = df_vis[df_vis['COLOR_NOMBRE'].isin(filtro_color)]

                cols_show = [c for c in cols_show_base if c in df_r.columns]
                st.dataframe(
                    df_vis[cols_show].style.format({
                        'PCT_LINEA': '{:.1%}', 'PCT_DISPO': '{:.1%}',
                        'LBS_C': '{:,.1f}', 'LBS_ASIGNADO': '{:,.1f}',
                        'LBS_FALTANTE': '{:,.1f}', 'INV_EFECTIVO': '{:,.0f}',
                    }),
                    use_container_width=True, hide_index=True, height=380
                )
                st.caption(f"{len(df_vis):,} líneas · {df_vis['DISPO'].nunique():,} dispos")

        # ── Descargar ─────────────────────────────────────────────────────────
        st.markdown('<p class="section-title">Descargar resultado</p>', unsafe_allow_html=True)
        ts = datetime.now().strftime('%Y%m%d%H%M')
        excel_bytes = generar_excel(resultados, entrega_pesos, cascada_activo, capacidad_cfg)
        st.download_button(
            label=f"⬇️  Descargar Excel — INVENTARIO_DISPOS_{ts}.xlsx",
            data=excel_bytes,
            file_name=f"INVENTARIO_DISPOS_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    else:
        st.markdown("""
        <div style="display:flex;align-items:center;justify-content:center;height:450px;
                    border:2px dashed #d1d5db;border-radius:12px;color:#9ca3af;
                    flex-direction:column;gap:0.5rem;">
            <div style="font-size:2.5rem">📂</div>
            <div style="font-weight:600;font-size:1rem;">Sube tu archivo y presiona Procesar</div>
            <div style="font-size:0.82rem;">Se calcularán los 3 escenarios simultáneamente</div>
        </div>
        """, unsafe_allow_html=True)
